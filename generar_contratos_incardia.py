#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Parser de "Informe Contratos Competencia.xlsx" -> panel_contratos_incardia.html

Reconstruido porque el script original (generar_panel.py, mencionado en el
README de "Solicitudes Paula") ya no existe en disco. El HTML resultante se
reconstruyó a partir del archivo panel_contratos_incardia.html ya publicado
(se extrajo como plantilla en panel_contratos_template.html, con los datos
reemplazados por placeholders).

Estructura del Excel (hoja de datos, no la de gráfico):
- Fila 1: encabezados -> Termino, Institución, Licitación, Descripción
  Licitación, Item, Descripción Item, Competidor, Adjudicado, Eurosets.
- Cada fila = una línea de item dentro de una licitación adjudicada a un
  competidor. "Competidor" es texto libre (razón social completa), se
  normaliza a una de las empresas monitoreadas por coincidencia de palabra
  clave (ver EMPRESA_KEYWORDS).
- "Eurosets" (columna I): si tiene texto -> compite Eurosets en ese item
  (EuroComp = "Sí"); si está vacío -> "No".

El Excel se descarga en vivo desde SharePoint vía Microsoft Graph API (ver
sharepoint.py) — no depende de ningún archivo local.

Variables de entorno requeridas:
- SP_TENANT_ID, SP_CLIENT_ID, SP_CLIENT_SECRET : credenciales del App
  Registration de Azure AD (las mismas que usa Panel Licitaciones).
- SP_FILE_URL_CONTRATOS : link de "Compartir" del Excel de contratos.

Uso:
  python generar_contratos_incardia.py
"""

import json
import os
from datetime import datetime
from pathlib import Path

import openpyxl

import sharepoint

BASE = Path(__file__).parent
TEMPLATE_PATH = BASE / "panel_contratos_template.html"
OUT_PATH = BASE / "panel_contratos_incardia.html"

SP_FILE_URL_CONTRATOS = os.environ["SP_FILE_URL_CONTRATOS"]

# Orden y color ya definidos en el HTML original - no cambiar el orden,
# el JS usa el mismo índice para COMP y COMP_COLORS.
EMPRESAS = ["Kaplan", "Medtronic", "Terumo", "Cardiotec", "Fresenius", "Gemco", "Incardia"]

# Palabras clave para reconocer cada empresa dentro de la razón social libre
# de la columna "Competidor". Case-insensitive, sin acentos.
EMPRESA_KEYWORDS = {
    "Kaplan": ["kaplan"],
    "Medtronic": ["medtronic"],
    "Terumo": ["terumo"],
    "Cardiotec": ["cardiotec"],
    "Fresenius": ["fresenius"],
    "Gemco": ["gemco"],
    "Incardia": ["incardia"],
}

MESES_EN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def normalizar(s):
    s = str(s or "").lower()
    s = (s.replace("á", "a").replace("é", "e").replace("í", "i")
          .replace("ó", "o").replace("ú", "u").replace("ñ", "n"))
    return s


def empresa_desde_competidor(texto):
    norm = normalizar(texto)
    for empresa in EMPRESAS:
        for kw in EMPRESA_KEYWORDS[empresa]:
            if kw in norm:
                return empresa
    return None


def encontrar_hoja_datos(wb):
    for name in wb.sheetnames:
        ws = wb[name]
        if hasattr(ws, "max_row") and hasattr(ws, "cell"):
            # Es una hoja de datos real (no un Chartsheet)
            return ws
    raise SystemExit("No se encontró una hoja de datos en el Excel.")


def parse_excel():
    excel_bytes = sharepoint.descargar_excel(
        SP_FILE_URL_CONTRATOS, "Informe Contratos Competencia.xlsx"
    )
    wb = openpyxl.load_workbook(excel_bytes, data_only=True)
    ws = encontrar_hoja_datos(wb)

    hoy = datetime.now()
    registros = []

    for r in range(2, ws.max_row + 1):
        termino = ws.cell(row=r, column=1).value
        institucion = ws.cell(row=r, column=2).value
        licitacion = ws.cell(row=r, column=3).value
        desc_lic = ws.cell(row=r, column=4).value
        desc_item = ws.cell(row=r, column=6).value
        competidor = ws.cell(row=r, column=7).value
        monto = ws.cell(row=r, column=8).value
        eurosets = ws.cell(row=r, column=9).value

        if not isinstance(termino, datetime) or not competidor or not licitacion:
            continue

        empresa = empresa_desde_competidor(competidor)
        if empresa is None:
            continue  # no es una de las empresas monitoreadas

        dias_restantes = (termino.date() - hoy.date()).days

        registros.append({
            "Licitación": str(licitacion).strip(),
            "Empresa": empresa,
            "Institución": str(institucion or "").strip(),
            "Termino_str": termino.strftime("%d/%m/%Y"),
            "DiasRestantes": dias_restantes,
            "Monto": float(monto) if monto else 0,
            "DescLic": str(desc_lic or "").strip(),
            "DescItem": str(desc_item or "").strip(),
            "EuroComp": "Sí" if (eurosets and str(eurosets).strip()) else "No",
            "_mes_key": (termino.year, termino.month),
        })

    return registros


def construir_monthly_data(registros):
    por_mes = {}
    for reg in registros:
        clave = reg["_mes_key"]
        por_mes.setdefault(clave, {emp: 0 for emp in EMPRESAS})
        por_mes[clave][reg["Empresa"]] += reg["Monto"]

    meses_ordenados = sorted(por_mes.keys())
    monthly = []
    for (anio, mes) in meses_ordenados:
        fila = {"mes": f"{MESES_EN[mes - 1]} {str(anio)[2:]}"}
        for emp in EMPRESAS:
            # en millones, redondeado, igual que el HTML original
            fila[emp] = round(por_mes[(anio, mes)][emp] / 1e6)
        monthly.append(fila)
    return monthly


def generar():
    registros = parse_excel()
    monthly = construir_monthly_data(registros)
    registros_limpios = [{k: v for k, v in r.items() if k != "_mes_key"} for r in registros]

    empresas_con_datos = sorted({r["Empresa"] for r in registros_limpios})

    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    html = template.replace("__ALL_CONTRACTS_JSON__", json.dumps(registros_limpios, ensure_ascii=False))
    html = html.replace("__MONTHLY_DATA_JSON__", json.dumps(monthly, ensure_ascii=False))

    faltantes = [e for e in EMPRESAS if e not in empresas_con_datos]
    nota_faltantes = f"{', '.join(faltantes)} sin contratos vigentes en esta corrida" if faltantes else ""
    footer_stats = f"{len(registros_limpios)} contratos &nbsp;&middot;&nbsp; {len(empresas_con_datos)} empresas monitoreadas"
    html = html.replace("__FOOTER_STATS__", footer_stats)

    OUT_PATH.parent.mkdir(exist_ok=True)
    OUT_PATH.write_text(html, encoding="utf-8")
    print(f"{len(registros_limpios)} contratos ({len(empresas_con_datos)} empresas) -> {OUT_PATH}")


if __name__ == "__main__":
    generar()
